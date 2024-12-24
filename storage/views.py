from django.shortcuts import render, redirect
from django.views.generic import ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import HttpResponse
from django.conf import settings
from .models import Stock, Warehouse
from gallery.models import SPU
from . import sync
import logging
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
from PIL import Image as PILImage
from io import BytesIO
import requests

logger = logging.getLogger(__name__)

class StockListView(LoginRequiredMixin, ListView):
    model = Stock
    template_name = 'storage/stock_list.html'
    context_object_name = 'stocks'
    paginate_by = 100
    login_url = '/muggle/login/'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        warehouse_id = self.request.GET.get('warehouse')
        product_type = self.request.GET.get('product_type')
        
        if search_query:
            queryset = queryset.filter(
                sku__sku_code__icontains=search_query
            )
        
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
            
        if product_type:
            queryset = queryset.filter(sku__spu__product_type=product_type)
            
        return queryset.select_related('warehouse', 'sku', 'sku__spu')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['active_menu'] = 'stock'
        context['warehouses'] = Warehouse.objects.all()
        context['selected_warehouse'] = self.request.GET.get('warehouse', '')
        
        # 添加产品类型数据
        product_types = []
        for pt in SPU.PRODUCT_TYPE_CHOICES:
            count = SPU.objects.filter(product_type=pt[0]).count()
            if count > 0:  # 只显示有数据的类型
                product_types.append({
                    'value': pt[0],
                    'label': pt[1],
                    'count': count
                })
        context['product_types'] = product_types
        context['selected_product_type'] = self.request.GET.get('product_type', '')
        
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get('download') == 'true':
            return self.download_stock_list(request)
        return super().get(request, *args, **kwargs)
    
    def download_stock_list(self, request):
        # 获取筛选后的数据
        queryset = self.get_queryset()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "库存列表"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 图片列
        ws.column_dimensions['B'].width = 20  # SKU编码
        ws.column_dimensions['C'].width = 40  # SKU名称
        ws.column_dimensions['D'].width = 15  # 产品类型
        ws.column_dimensions['E'].width = 15  # 所属仓库
        ws.column_dimensions['F'].width = 12  # 库存数量
        ws.column_dimensions['G'].width = 12  # 平均成本
        ws.column_dimensions['H'].width = 20  # 更新时间
        
        # 设置表头样式
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        header_font = Font(bold=True)
        
        # 写入表头
        headers = ['图片', 'SKU编码', 'SKU名称', '产品类型', '所属仓库', '库存数量', '平均成本', '更新时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 设置行高
        ws.row_dimensions[1].height = 20
        
        # 写入数据
        for row_idx, stock in enumerate(queryset, 2):
            # 设置图片行高
            ws.row_dimensions[row_idx].height = 60
            
            # 处理图片
            if stock.sku.img_url:
                try:
                    img_path = os.path.join(settings.MEDIA_ROOT, str(stock.sku.img_url))
                    if os.path.exists(img_path):
                        # 使用 PIL 处理图片
                        pil_image = PILImage.open(img_path)
                        # 调整图片大小
                        pil_image.thumbnail((60, 60))
                        # 保存到 BytesIO
                        img_byte_arr = BytesIO()
                        pil_image.save(img_byte_arr, format=pil_image.format or 'PNG')
                        img_byte_arr.seek(0)
                        
                        # 创建 openpyxl 图片对象
                        img = Image(img_byte_arr)
                        # 调整图片大小
                        img.width = 60
                        img.height = 60
                        # 添加到单元格
                        ws.add_image(img, f'A{row_idx}')
                except Exception as e:
                    logger.error(f"处理图片失败: {str(e)}")
            
            # 写入其他数据
            ws.cell(row=row_idx, column=2, value=stock.sku.sku_code)
            ws.cell(row=row_idx, column=3, value=stock.sku.sku_name)
            ws.cell(row=row_idx, column=4, value=stock.sku.spu.get_product_type_display())
            ws.cell(row=row_idx, column=5, value=stock.warehouse.name)
            ws.cell(row=row_idx, column=6, value=stock.stock_num)
            ws.cell(row=row_idx, column=7, value=float(stock.avg_cost))
            ws.cell(row=row_idx, column=8, value=stock.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # 设置单元格对齐方式
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置表头居中对齐
        for col in range(1, 9):
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="stock_list_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # 保存工作簿
        wb.save(response)
        
        return response

class StockSyncView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            sync.sync_all_stock()
            messages.success(request, '库存数据同步成功！')
            logger.info(f"用户 {request.user.username} 同步库存数据成功")
        except Exception as e:
            error_msg = f'同步失败：{str(e)}'
            messages.error(request, error_msg)
            logger.error(f"用户 {request.user.username} 同步库存数据失败: {str(e)}")
        return redirect('storage:stock_list')
    def get(self, request, *args, **kwargs):
        try:
            sync.sync_all_stock()
            messages.success(request, '库存数据同步成功！')
            logger.info(f"用户 {request.user.username} 同步库存数据成功")
        except Exception as e:
            error_msg = f'同步失败：{str(e)}'
            messages.error(request, error_msg)
            logger.error(f"用户 {request.user.username} 同步库存数据失败: {str(e)}")
        return redirect('storage:stock_list')